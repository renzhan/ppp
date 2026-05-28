"""
生成字段映射与获取方式文档
读取 docs/业务底表与数据源字段映射关系.xlsx，为每一行二级模块输出值生成获取方式（SQL或其他）
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

# ── 读取源文件 ──
wb = openpyxl.load_workbook(r'docs/业务底表与数据源字段映射关系.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# ── 解析合并单元格，填充空值 ──
# 先记录所有合并区域的值
merge_map = {}
for merge_range in ws.merged_cells.ranges:
    min_row = merge_range.min_row
    min_col = merge_range.min_col
    val = ws.cell(min_row, min_col).value
    for row in range(merge_range.min_row, merge_range.max_row + 1):
        for col in range(merge_range.min_col, merge_range.max_col + 1):
            if row != min_row or col != min_col:
                merge_map[(row, col)] = val

def get_cell_value(row_idx, col_idx):
    """获取单元格值，考虑合并单元格"""
    val = ws.cell(row_idx, col_idx).value
    if val is not None:
        return str(val).strip()
    if (row_idx, col_idx) in merge_map:
        v = merge_map[(row_idx, col_idx)]
        return str(v).strip() if v is not None else ''
    return ''

# ── 数据库表字段映射 ──
# 根据 prisma schema 和代码整理

TABLE_FIELD_MAP = {
    # projects 表
    'projects': {
        '品类': 'category',
        '合作品牌': 'brand',
        '品牌业务线': 'business_line',
        '项目名称': 'project_name',
        '项目类型': 'project_type',
        '投放开始时间': 'start_date',
        '投放结束时间': 'end_date',
        'SPU名称': 'spu_name',
    },
    # note_base 表 (业务底表)
    'note_base': {
        '笔记id': 'note_id',
        '笔记链接': 'note_link',
        '合作形式': 'cooperation_form',
        '是否报备': 'is_registered',
        '内容方向': 'content_direction',
        '达人类型': 'kol_type',
        '对应SPU': 'spu_name',
        '内容实际消耗金额': 'content_cost',
        '内容实际结算金额': 'content_settlement',
        '投流实际消耗': 'ad_spend',
        '总费用': 'total_cost',
    },
    # notes 表 (蒲公英)
    'notes': {
        '曝光量': 'imp_num',
        '阅读量': 'read_num',
        '点赞量': 'like_num',
        '收藏量': 'fav_num',
        '评论量': 'cmt_num',
        '分享量': 'share_num',
        '关注量': 'follow_num',
        '博主粉丝量': 'kol_fan_num',
        '博主昵称': 'kol_nick_name',
        '笔记标题': 'note_title',
        '笔记类型': 'note_type',
        '笔记链接': 'note_link',
        '互动量': 'engage_num',
    },
    # juguang_data 表 (聚光)
    'juguang_data': {
        '消费': 'fee',
        '展现量': 'impression',
        '点击量': 'click',
        '互动量': 'interaction',
        '新增种草人群': 'i_user_num',
        '新增深度种草人群': 'ti_user_num',
        '新增种草人群成本': 'i_user_price',
        '新增深度种草人群成本': 'ti_user_price',
        '搜索组件点击量': 'search_cmt_click',
        '搜后阅读量': 'search_cmt_after_read',
    },
    # lingxi_data 表 (灵犀)
    'lingxi_data': {
        'AIPS人群总数': 'data_content->aips',
        'TI深度兴趣人群数': 'data_content->ti',
        '月搜索指数': 'data_content->monthlySearchVolume',
        '人群渗透率': 'data_content->penetrationRate',
    },
    # review_configs 表
    'review_configs': {
        '大盘': 'benchmark',
        'KPI目标': 'kpi_targets',
        '互动率口径': 'engagement_metric',
    },
    # comments 表
    'comments': {
        '评论内容': 'content',
        '评论时间': 'comment_time',
        '点赞数': 'likes',
    },
    # qiangua_data 表
    'qiangua_data': {
        '千瓜数据': 'data_content',
    },
}


def determine_acquisition_method(row_idx, one_level, two_level, output_val, formula, source_val, source_platform, api_info, remark):
    """
    根据来源平台和计算公式确定获取方式
    返回: (获取方式类型, 具体SQL或说明)
    """
    platform = source_platform.strip() if source_platform else ''
    formula_str = formula.strip() if formula else ''
    source = source_val.strip() if source_val else ''
    output = output_val.strip() if output_val else ''
    
    # AI解读
    if platform == 'AI解读' or 'AI' in formula_str:
        return ('AI生成', f'基于相关数据由LLM生成，调用LLM API（配置中的QWEN/OpenAI）生成文本')
    
    # 人工录入
    if platform == '人工录入':
        if '大盘' in output:
            return ('SQL查询', f"SELECT benchmark->'{output.replace('大盘-', '').lower()}' FROM review_configs WHERE project_id = '{{project_id}}' ORDER BY created_at DESC LIMIT 1")
        if 'KPI' in output or '目标' in output:
            return ('SQL查询', f"SELECT kpi_targets->'{output}' FROM review_configs WHERE project_id = '{{project_id}}' ORDER BY created_at DESC LIMIT 1")
        if '互动率' in output and '口径' not in output:
            return ('SQL查询', f"SELECT engagement_metric FROM review_configs WHERE project_id = '{{project_id}}' ORDER BY created_at DESC LIMIT 1")
        return ('人工录入', f"用户在「开始复盘」配置页面手动录入，存储在 review_configs 表的对应JSON字段中")
    
    # 系统库 (projects表)
    if platform == '系统库':
        field = TABLE_FIELD_MAP['projects'].get(output, None)
        if field:
            return ('SQL查询', f"SELECT {field} FROM projects WHERE id = '{{project_id}}'")
        return ('SQL查询', f"SELECT * FROM projects WHERE id = '{{project_id}}' -- 字段: {output}")
    
    # 蒲公英
    if platform == '蒲公英' or '蒲公英' in platform:
        if source in TABLE_FIELD_MAP['notes']:
            db_field = TABLE_FIELD_MAP['notes'][source]
            if 'SUM' in formula_str or 'GROUP BY' in formula_str:
                return ('SQL查询', f"SELECT SUM({db_field}) FROM notes WHERE project_id = '{{project_id}}'")
            if 'COUNT' in formula_str or '数量' in output or '篇数' in output:
                return ('SQL查询', f"SELECT COUNT(*) FROM notes WHERE project_id = '{{project_id}}'")
            return ('SQL查询', f"SELECT {db_field} FROM notes WHERE project_id = '{{project_id}}'")
        if 'note_title' in source:
            return ('SQL查询', f"SELECT note_title FROM notes WHERE project_id = '{{project_id}}'")
        if 'kol_nick_name' in source:
            return ('SQL查询', f"SELECT kol_nick_name FROM notes WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
        if 'like_num' in source:
            return ('SQL查询', f"SELECT like_num FROM notes WHERE project_id = '{{project_id}}'")
        if 'fav_num' in source:
            return ('SQL查询', f"SELECT fav_num FROM notes WHERE project_id = '{{project_id}}'")
        if 'cmt_num' in source:
            return ('SQL查询', f"SELECT cmt_num FROM notes WHERE project_id = '{{project_id}}'")
        if 'share_num' in source:
            return ('SQL查询', f"SELECT share_num FROM notes WHERE project_id = '{{project_id}}'")
        # 笔记媒体采集
        if '笔记媒体' in str(api_info):
            return ('外部API', f"调用笔记媒体采集API: POST http://47.114.122.60:6002/api/note-media/collect，传入note_id列表")
        return ('SQL查询', f"SELECT * FROM notes WHERE project_id = '{{project_id}}' -- 来源值: {source}")
    
    # 聚光
    if platform == '聚光' or '聚光' in platform:
        # 聚光数据存在 juguang_data 表
        if source in TABLE_FIELD_MAP['juguang_data']:
            db_field = TABLE_FIELD_MAP['juguang_data'][source]
        elif source == 'fee':
            db_field = 'fee'
        elif source == 'impression':
            db_field = 'impression'
        elif source == 'click':
            db_field = 'click'
        elif source == 'interaction':
            db_field = 'interaction'
        elif source == 'i_user_num':
            db_field = 'i_user_num'
        elif source == 'ti_user_num':
            db_field = 'ti_user_num'
        elif source == 'i_user_price':
            db_field = 'i_user_price'
        elif source == 'ti_user_price':
            db_field = 'ti_user_price'
        elif source == 'placement':
            db_field = 'placement'
        elif source == 'targetDetail':
            db_field = 'target_detail'
        elif source == 'keyword':
            db_field = 'keyword'
        else:
            db_field = source
        
        # 解析 GROUP BY 公式生成 SQL
        if 'GROUP BY' in formula_str:
            # 提取分组维度和聚合
            if 'SUM(fee)/SUM(impression)*1000' in formula_str:
                return ('SQL查询', f"SELECT SUM(fee)::numeric / NULLIF(SUM(impression),0) * 1000 AS cpm FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(fee)/SUM(click)' in formula_str:
                return ('SQL查询', f"SELECT SUM(fee)::numeric / NULLIF(SUM(click),0) AS cpc FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(fee)/SUM(interaction)' in formula_str:
                return ('SQL查询', f"SELECT SUM(fee)::numeric / NULLIF(SUM(interaction),0) AS cpe FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(click)/SUM(impression)*100' in formula_str:
                return ('SQL查询', f"SELECT SUM(click)::numeric / NULLIF(SUM(impression),0) * 100 AS ctr FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(fee)/SUM(i_user_num)' in formula_str:
                return ('SQL查询', f"SELECT SUM(fee)::numeric / NULLIF(SUM(i_user_num),0) AS i_user_price FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(fee)/SUM(ti_user_num)' in formula_str:
                return ('SQL查询', f"SELECT SUM(fee)::numeric / NULLIF(SUM(ti_user_num),0) AS ti_user_price FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(fee)' in formula_str:
                return ('SQL查询', f"SELECT SUM(fee) FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(impression)' in formula_str:
                return ('SQL查询', f"SELECT SUM(impression) FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(click)' in formula_str:
                return ('SQL查询', f"SELECT SUM(click) FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(interaction)' in formula_str:
                return ('SQL查询', f"SELECT SUM(interaction) FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(i_user_num)' in formula_str:
                return ('SQL查询', f"SELECT SUM(i_user_num) FROM juguang_data WHERE project_id = '{{project_id}}'")
            elif 'SUM(ti_user_num)' in formula_str:
                return ('SQL查询', f"SELECT SUM(ti_user_num) FROM juguang_data WHERE project_id = '{{project_id}}'")
            else:
                return ('SQL查询', f"SELECT {db_field} FROM juguang_data WHERE project_id = '{{project_id}}' -- 公式: {formula_str}")
        
        # 聚光关键词规划工具
        if '关键词规划工具' in str(api_info) or '关键词规划' in str(remark):
            return ('SQL缺失字段', f"juguang_data表当前无此字段存储，需调用聚光API: 关键词规划工具接口获取相关词列表")
        
        return ('SQL查询', f"SELECT {db_field} FROM juguang_data WHERE project_id = '{{project_id}}'")
    
    # 灵犀
    if platform == '灵犀' or '灵犀' in platform:
        return ('SQL查询', f"SELECT data_type, data_content FROM lingxi_data WHERE project_id = '{{project_id}}' AND data_type = 'brand' -- 具体字段从data_content JSON中提取")
    
    # 千瓜
    if platform == '千瓜' or '千瓜' in platform:
        return ('SQL查询', f"SELECT data_type, data_content FROM qiangua_data WHERE project_id = '{{project_id}}'")
    
    # 执行业务底表
    if '业务底表' in platform or '执行业务底表' in platform or '底表' in platform:
        if '内容方向' in source or '内容方向' in output:
            return ('SQL查询', f"SELECT content_direction FROM note_base WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
        if '达人类型' in source or '达人类型' in output:
            return ('SQL查询', f"SELECT kol_type FROM note_base WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
        if '合作形式' in source or '合作形式' in output:
            return ('SQL查询', f"SELECT cooperation_form FROM note_base WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
        if '词类' in source or '词类' in output:
            return ('SQL缺失字段', f"note_base表当前无「词类」字段，需新增word_category字段或从业务标注中获取")
        # COUNT类 — 必须在博主主页链接检查之前，因为source可能是博主主页链接但output是数量
        if '数量' in output or '篇数' in output:
            return ('SQL查询', f"SELECT COUNT(*) FROM note_base WHERE project_id = '{{project_id}}' -- 原公式COUNT(博主主页链接)，用COUNT(note_id)替代")
        # 爆文率等需要关联notes表
        if '爆文' in output:
            return ('SQL查询', f"SELECT COUNT(*) FILTER (WHERE n.like_num+n.fav_num+n.cmt_num+n.share_num >= 1000) AS hot_count, COUNT(*) AS total FROM notes n WHERE n.project_id = '{{project_id}}' -- 爆文阈值根据review_configs配置确定")
        if '博主主页链接' in source or '博主主页链接' in output:
            return ('SQL缺失字段', f"note_base表当前无「博主主页链接」字段，需新增kol_home_link字段")
        if '内容实际消耗' in source or '内容消耗' in output:
            return ('SQL查询', f"SELECT content_cost FROM note_base WHERE project_id = '{{project_id}}'")
        if '内容实际结算' in source or '内容结算' in output:
            return ('SQL查询', f"SELECT content_settlement FROM note_base WHERE project_id = '{{project_id}}'")
        if '投流实际消耗' in source or '投流消耗' in output:
            return ('SQL查询', f"SELECT ad_spend FROM note_base WHERE project_id = '{{project_id}}'")
        if '总费用' in source or '总费用' in output:
            return ('SQL查询', f"SELECT total_cost FROM note_base WHERE project_id = '{{project_id}}'")
        if '笔记id' in source.lower() or 'note_id' in source.lower():
            return ('SQL查询', f"SELECT note_id FROM note_base WHERE project_id = '{{project_id}}'")
        if '是否报备' in source or '是否报备' in output:
            return ('SQL查询', f"SELECT is_registered FROM note_base WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
        if 'SPU' in source or 'spu' in source.lower() or 'SPU' in output:
            return ('SQL查询', f"SELECT spu_name FROM note_base WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
        return ('SQL查询', f"SELECT * FROM note_base WHERE project_id = '{{project_id}}' AND note_id = '{{note_id}}'")
    
    # 系统计算
    if '系统计算' in platform or platform == '系统计算':
        if 'KPI' in output or '完成率' in output:
            return ('系统计算', f"系统自动对比实际达成值与KPI目标值计算完成率: 实际值/目标值*100%")
        if '大盘' in output or '优于大盘' in output:
            return ('系统计算', f"系统自动对比实际达成值与review_configs.benchmark中的大盘均值")
        return ('系统计算', f"基于已有数据由系统自动计算: {formula_str}")
    
    # 舆情/评论
    if '舆情' in platform or '评论' in platform:
        return ('SQL查询', f"SELECT * FROM comments WHERE project_id = '{{project_id}}'")
    
    # 灵犀截图OCR
    if '灵犀（截图OCR）' in platform or 'OCR' in platform:
        return ('无法确定', f"待开发功能：用户上传灵犀截图，通过OCR+AI进行解读，目前无结构化API")
    
    # 派查查
    if '派查查' in platform:
        return ('无法确定', f"派查查平台数据，当前系统未对接此数据源的结构化存储")
    
    # 空或未知
    if not platform and not formula_str:
        # 尝试从其他字段推断
        if source and '蒲公英' in source:
            return ('SQL查询', f"SELECT * FROM notes WHERE project_id = '{{project_id}}' -- 来源: {source}")
        if source and '灵犀' in source:
            return ('SQL查询', f"SELECT data_content FROM lingxi_data WHERE project_id = '{{project_id}}' -- 来源: {source}")
        if source and '聚光' in source:
            return ('SQL查询', f"SELECT * FROM juguang_data WHERE project_id = '{{project_id}}' -- 来源: {source}")
        return ('无法确定', f'来源平台和计算公式均为空')
    
    if not platform and formula_str:
        # 公式中有平台信息
        if '蒲公英' in formula_str:
            if source and '蒲公英' in source:
                return ('SQL查询', f"SELECT * FROM notes WHERE project_id = '{{project_id}}' -- 数据来自蒲公英API爬取")
            return ('SQL查询', f"SELECT * FROM notes WHERE project_id = '{{project_id}}' -- 公式提及蒲公英")
        if '灵犀' in formula_str:
            return ('SQL查询', f"SELECT data_content FROM lingxi_data WHERE project_id = '{{project_id}}' -- 公式提及灵犀")
        if '聚光' in formula_str:
            return ('SQL查询', f"SELECT * FROM juguang_data WHERE project_id = '{{project_id}}' -- 公式提及聚光")
        if '业务底表' in formula_str:
            if '博主主页链接' in output or '博主主页链接' in source:
                return ('SQL缺失字段', f"note_base表当前无「博主主页链接」字段，需新增kol_home_link字段")
            return ('SQL查询', f"SELECT * FROM note_base WHERE project_id = '{{project_id}}' -- 公式提及业务底表")
        if '截止' in formula_str or '时间范围' in formula_str or '类目层级' in formula_str:
            return ('参数说明', f"此行为上一行数据获取的补充参数说明: {formula_str}")
        if '=聚光API' in formula_str or '=灵犀API' in formula_str:
            if '聚光' in formula_str or '聚光' in source:
                return ('SQL缺失字段', f"需调用聚光外部API获取，当前juguang_data表可能无此字段: {source}")
            return ('SQL查询', f"SELECT data_content FROM lingxi_data WHERE project_id = '{{project_id}}' -- {source}")
        # 公式本身是解读文本模板
        if len(formula_str) > 50:
            return ('AI生成', f"此行为AI解读的模板/示例文本，由LLM基于数据生成")
        return ('无法确定', f'来源平台为空, 公式: {formula_str}, 来源值: {source}')

    if '直接引用' in formula_str:
        if source and ('蒲公英' in source or 'like_num' in source or 'fav_num' in source or 'cmt_num' in source):
            return ('SQL查询', f"SELECT * FROM notes WHERE project_id = '{{project_id}}' -- 引用蒲公英数据")
        return ('无法确定', f'标注为直接引用但未明确数据来源表: {source}')
    
    return ('无法确定', f'来源平台: {platform}, 公式: {formula_str}, 来源值: {source}')


def determine_source_table(platform, source_val, output_val, formula, acq_detail):
    """根据来源平台和SQL内容确定来源数据库表"""
    tables = set()
    platform = platform or ''
    source_val = source_val or ''
    output_val = output_val or ''
    formula = formula or ''
    acq_detail = acq_detail or ''
    
    # 从SQL中提取表名
    for tbl in ['projects', 'note_base', 'notes', 'lingxi_data', 'juguang_data', 'qiangua_data', 'comments', 'review_configs']:
        if tbl in acq_detail:
            tables.add(tbl)
    
    # 从来源平台推断
    if '系统库' in platform:
        tables.add('projects')
    if '蒲公英' in platform:
        tables.add('notes')
    if '聚光' in platform:
        tables.add('juguang_data')
    if '灵犀' in platform:
        tables.add('lingxi_data')
    if '千瓜' in platform:
        tables.add('qiangua_data')
    if '业务底表' in platform or '执行业务底表' in platform:
        tables.add('note_base')
    if '人工录入' in platform:
        tables.add('review_configs')
    if '舆情' in platform or '评论' in platform:
        tables.add('comments')
    
    # 从公式/来源值补充推断
    if '蒲公英' in formula or '蒲公英' in source_val:
        tables.add('notes')
    if '聚光' in formula or '聚光' in source_val:
        tables.add('juguang_data')
    if '灵犀' in formula or '灵犀' in source_val:
        tables.add('lingxi_data')
    if '业务底表' in formula or '底表' in source_val:
        tables.add('note_base')
    
    # 交叉分析场景（需要多表关联）
    if '内容方向' in formula and ('聚光' in platform or 'juguang' in acq_detail):
        tables.add('note_base')  # 内容方向来自底表
        tables.add('juguang_data')
    if '创作者' in formula and '聚光' in platform:
        tables.add('notes')  # 博主昵称来自蒲公英
        tables.add('juguang_data')
    
    if not tables:
        return '—'
    return ', '.join(sorted(tables))


# ── 读取所有行并生成结果 ──
results = []
current_one_level = ''
current_two_level = ''

for row_idx in range(2, ws.max_row + 1):
    one_level = get_cell_value(row_idx, 1)  # A: 一级模块
    show_one = get_cell_value(row_idx, 2)   # B: 报告是否一定展示
    two_level = get_cell_value(row_idx, 3)  # C: 二级模块
    show_two = get_cell_value(row_idx, 4)   # D: 报告是否一定展示
    output_val = get_cell_value(row_idx, 5) # E: 输出值
    formula = get_cell_value(row_idx, 6)    # F: 计算公式
    source_val = get_cell_value(row_idx, 7) # G: 来源值
    source_platform = get_cell_value(row_idx, 8)  # H: 来源平台
    api_info = get_cell_value(row_idx, 9)   # I: 接口
    remark = get_cell_value(row_idx, 10)    # J: 备注
    
    # 跟踪当前一级/二级模块
    if one_level:
        current_one_level = one_level
    if two_level:
        current_two_level = two_level
    
    # 跳过完全空行
    if not output_val and not formula and not source_val and not source_platform:
        continue
    
    # 确定获取方式
    acq_type, acq_detail = determine_acquisition_method(
        row_idx, current_one_level, current_two_level,
        output_val, formula, source_val, source_platform, api_info, remark
    )
    
    # 确定来源表
    source_table = determine_source_table(source_platform, source_val, output_val, formula, acq_detail)
    
    results.append({
        'row': row_idx,
        'one_level': current_one_level,
        'two_level': current_two_level,
        'output_val': output_val,
        'formula': formula,
        'source_val': source_val,
        'source_platform': source_platform,
        'api_info': api_info,
        'remark': remark,
        'acq_type': acq_type,
        'acq_detail': acq_detail,
        'source_table': source_table,
    })

# ── 生成组合SQL（按一级模块/章节+来源表分组） ──
def generate_combined_sql(all_results):
    """
    按一级模块（章节）分组，同一章节内来源表相同的行合并为一条SQL，
    SELECT 所有需要的字段。每行都写同一个组合SQL，方便按值筛选去重。
    """
    from collections import defaultdict
    
    # 按 (一级模块, 来源表) 分组收集字段
    groups = defaultdict(lambda: {'fields': [], 'has_agg': False})
    
    for r in all_results:
        if r['acq_type'] != 'SQL查询':
            continue
        tables_str = r['source_table']
        if tables_str == '—' or not tables_str:
            continue
        
        key = (r['one_level'], tables_str)
        detail = r['acq_detail']
        
        # 从SQL中提取字段
        if 'SELECT ' in detail:
            try:
                select_part = detail.split('SELECT ')[1].split(' FROM ')[0].strip()
                # 跳过 * 
                if select_part == '*':
                    if '*' not in groups[key]['fields']:
                        groups[key]['fields'].append('*')
                else:
                    # 检查是否有聚合函数
                    if 'SUM(' in select_part or 'COUNT(' in select_part or 'AVG(' in select_part:
                        groups[key]['has_agg'] = True
                    if select_part not in groups[key]['fields']:
                        groups[key]['fields'].append(select_part)
            except (IndexError, ValueError):
                pass
    
    # 为每个分组生成组合SQL
    combined = {}
    for (one_level, tables_str), info in groups.items():
        fields = info['fields']
        if not fields:
            continue
        
        # 如果有 * 就直接用 *
        if '*' in fields:
            select_fields = '*'
        else:
            select_fields = ', '.join(fields)
        
        # 确定主表
        tables = [t.strip() for t in tables_str.split(',')]
        main_table = tables[0].strip()
        
        # 构建SQL
        if len(tables) == 1:
            sql = f"SELECT {select_fields} FROM {main_table} WHERE project_id = '{{project_id}}'"
        else:
            # 多表JOIN
            alias_map = {
                'notes': 'n',
                'note_base': 'nb', 
                'juguang_data': 'jd',
                'lingxi_data': 'ld',
                'qiangua_data': 'qd',
                'comments': 'c',
                'review_configs': 'rc',
                'projects': 'p',
            }
            main_alias = alias_map.get(main_table, main_table[0])
            joins = []
            for t in tables[1:]:
                t = t.strip()
                a = alias_map.get(t, t[0])
                if t in ('notes', 'note_base', 'juguang_data'):
                    joins.append(f"LEFT JOIN {t} {a} ON {a}.project_id = {main_alias}.project_id AND {a}.note_id = {main_alias}.note_id")
                else:
                    joins.append(f"LEFT JOIN {t} {a} ON {a}.project_id = {main_alias}.project_id")
            join_str = ' '.join(joins)
            sql = f"SELECT {select_fields} FROM {main_table} {main_alias} {join_str} WHERE {main_alias}.project_id = '{{project_id}}'"
        
        combined[(one_level, tables_str)] = sql
    
    return combined

combined_sql_map = generate_combined_sql(results)

# 为每行分配组合SQL
for r in results:
    key = (r['one_level'], r['source_table'])
    if r['acq_type'] == 'SQL查询' and key in combined_sql_map:
        r['combined_sql'] = combined_sql_map[key]
    else:
        r['combined_sql'] = ''

# ── 写入输出 Excel ──
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = '字段获取方式'

# 表头
headers = ['序号', '一级模块', '二级模块', '输出值', '计算公式', '来源值', '来源平台', '来源表', '获取方式类型', '具体获取方式/SQL', '组合SQL']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = out_ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 数据行
type_colors = {
    'SQL查询': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    'AI生成': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    '人工录入': PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    '系统计算': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    '外部API': PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
    'SQL缺失字段': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    '无法确定': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
}

for i, r in enumerate(results, 1):
    row_num = i + 1
    out_ws.cell(row=row_num, column=1, value=i)
    out_ws.cell(row=row_num, column=2, value=r['one_level'])
    out_ws.cell(row=row_num, column=3, value=r['two_level'])
    out_ws.cell(row=row_num, column=4, value=r['output_val'])
    out_ws.cell(row=row_num, column=5, value=r['formula'])
    out_ws.cell(row=row_num, column=6, value=r['source_val'])
    out_ws.cell(row=row_num, column=7, value=r['source_platform'])
    out_ws.cell(row=row_num, column=8, value=r['source_table'])
    out_ws.cell(row=row_num, column=9, value=r['acq_type'])
    out_ws.cell(row=row_num, column=10, value=r['acq_detail'])
    out_ws.cell(row=row_num, column=11, value=r['combined_sql'])
    
    # 设置获取方式类型的颜色
    type_cell = out_ws.cell(row=row_num, column=9)
    if r['acq_type'] in type_colors:
        type_cell.fill = type_colors[r['acq_type']]
    
    # 设置边框和自动换行
    for col in range(1, 12):
        cell = out_ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# 设置列宽
col_widths = [6, 14, 22, 22, 40, 16, 12, 20, 14, 80, 100]
for i, w in enumerate(col_widths, 1):
    out_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# 保存
output_path = r'docs/字段获取方式映射表.xlsx'
out_wb.save(output_path)
print(f'已生成: {output_path}')
print(f'总行数: {len(results)}')

# 统计
from collections import Counter
type_counts = Counter(r['acq_type'] for r in results)
print('\n获取方式统计:')
for t, c in type_counts.most_common():
    print(f'  {t}: {c}')
